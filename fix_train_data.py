import re
from pprint import pprint

import scapy.all as sc
from scapy.utils import rdpcap
# 打开一张表
from openpyxl import Workbook, load_workbook

from utils import *


def read_txt(filename):
    """
    读取txt
    :param filename: 文件名
    :return:
    """
    info_list = list()
    start_line = 3
    end_line = 33
    pattern = r'(.+?)\s+([\w]{2}:[\w]{2}:[\w]{2}:[\w]{2}:[\w]{2}:[\w]{2})'
    with open(filename, "r") as f:
        line_num = 0
        for line in f.readlines():
            line_num += 1
            if line_num < start_line:
                continue
            elif line_num > end_line:
                break
            else:
                line = line.strip("\n")
                search_obj = re.search(pattern, line)
                info_list.append((search_obj.group(1), search_obj.group(2)))
    return info_list


def get_eths_info():
    """
    获取mac地址
    :return:
    """
    info_list = read_txt("train_data/List_Of_Devices.txt")
    eths_info = dict()
    for info in info_list:
        eths_info[info[1]] = {"device": info[0],
                              "ip": None}
    count = 0
    pkts = rdpcap(TRAIN_PCAP_FILE)
    for i, pkt in enumerate(pkts):
        # print("packet index: {packet_index}".format(packet_index=i+1))
        ethernet = pkt.dst
        if eths_info[ethernet]["ip"] is None:
            if sc.IP in pkt:
                eths_info[ethernet]["ip"] = pkt[sc.IP].dst
            else:
                eths_info[ethernet]["ip"] = pkt[sc.IPv6].dst
            count += 1
            if count == 31:
                break
    device_count = 0
    for eth, eht_info in eths_info.items():
        if eht_info["ip"] is not None:
            device_count += 1
    print(device_count)
    pprint(eths_info)
    return eths_info


def write_xlsx(eths_info):
    """
    将对应信息写到xlsx中
    :param eths_info: mac信息
    :return:
    """
    res_wb = Workbook()
    res_ws = res_wb.active
    row = 1
    device_column = 1
    type_column = 2
    vendor_column = 3
    ip_column = 4
    ethernet_column = 5
    res_ws.cell(row, device_column).value = "device"
    res_ws.cell(row, type_column).value = "type"
    res_ws.cell(row, vendor_column).value = "vendor"
    res_ws.cell(row, ip_column).value = "ip"
    res_ws.cell(row, ethernet_column).value = "ethernet"
    for eth, eth_info in eths_info.items():
        row += 1
        res_ws.cell(row, device_column).value = eth_info["device"]
        res_ws.cell(row, ip_column).value = eth_info["ip"]
        res_ws.cell(row, ethernet_column).value = eth
    res_wb.save(os.path.join(TRAIN_DATA_FOLDER_NAME, TRAIN_TARGET_NAME + ".xlsx"))


def sort_xlsx():
    """
    对xlsx排序
    :return:
    """
    raw_wb = load_workbook(os.path.join(TRAIN_DATA_FOLDER_NAME, TRAIN_TARGET_NAME + ".xlsx"))
    raw_ws = raw_wb.active
    device_column = 1
    type_column = 2
    vendor_column = 3
    ip_column = 4
    ethernet_column = 5
    res_wb = Workbook()
    res_ws = res_wb.active
    start_row = 2
    end_row = 32
    ips_info_dict = dict()
    for row in range(start_row, end_row + 1):
        ip = raw_ws.cell(row, ip_column).value
        device = raw_ws.cell(row, device_column).value
        ethernet = raw_ws.cell(row, ethernet_column).value
        ips_info_dict[ip] = {"device": device, "ethernet": ethernet}
    ips_info_dict = get_sorted_dict_by_ip(ips_info_dict)
    row = 1
    res_ws.cell(row, device_column).value = "device"
    res_ws.cell(row, type_column).value = "type"
    res_ws.cell(row, vendor_column).value = "vendor"
    res_ws.cell(row, ip_column).value = "ip"
    res_ws.cell(row, ethernet_column).value = "ethernet"
    for ip, ip_info in ips_info_dict.items():
        row += 1
        res_ws.cell(row, device_column).value = ip_info["device"]
        res_ws.cell(row, ethernet_column).value = ip_info["ethernet"]
        res_ws.cell(row, ip_column).value = ip
    res_wb.save(os.path.join(TRAIN_DATA_FOLDER_NAME, TRAIN_TARGET_NAME + "_new.xlsx"))


def get_train_ips_device_info():
    """
    从excel中获取训练ip的信息 ip: {"device": device, "type": type, "vendor": vendor}
    :return:
    """
    raw_wb = load_workbook(TRAIN_IPS_DEVICE_INFO_XLSX)
    # 从表单中得到单元格内容
    raw_ws = raw_wb.active
    device_name_col = 1
    type_col = 2
    vendor_col = 3
    ip_col = 4
    ethernet_col = 5
    start_row = 2
    end_row = 67
    ips_devices_info = dict()
    for row in range(start_row, end_row + 1):
        vendor = raw_ws.cell(row, vendor_col).value
        if vendor == "non-iot":
            continue
        ip = raw_ws.cell(row, ip_col).value
        if ip not in BLACK_IPS:
            device = raw_ws.cell(row, device_name_col).value
            type = raw_ws.cell(row, type_col).value
            vendor = raw_ws.cell(row, vendor_col).value
            ips_devices_info[ip] = {"device": device, "type": type, "vendor": vendor}
    store_json(ips_devices_info, TRAIN_IPS_DEVICE_INFO_FILE)


def main():
    # info_list = read_txt("train_data/List_Of_Devices.txt")
    # print(info_list)
    # eths_info = get_eths_info()
    # write_xlsx(eths_info)
    # sort_xlsx()
    get_train_ips_device_info()


if __name__ == '__main__':
    main()
