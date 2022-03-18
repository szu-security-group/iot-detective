import os
import numpy as np

from constants import *
# 打开一张表
from openpyxl import Workbook

# 常量
DEVICE_TO_IP_FILE = "res/device_to_ip.json"
DEVICES_INFO_FILE = "res/devices_info.json"  # devices: [type, vendor, ip]
DEVICES_GREAT_DOMAINS_TO_DEVICES_FILE = "res/devices_great_domains_to_devices.json"
DEVICES_DOMAINS_DETECTION_FILE = "res/devices_domains_detection.json"
DEVICES_DOMAINS_TFIDF_VECTOR_LENGTH_FILE = "res/devices_domains_tfidf_vector_length.json"
DEVICES_DOMAINS_TFIDF_FILE = "res/devices_domains_tfidf.json"
DEVICES_MIXED_DOMAINS_TFIDF_FILE = "res/devices_mixed_domains_tfidf.json"
DOMAINS_WHOIS_INFO_FILE = "res/domains_whois_info.json"
HIDDEN_ORGANIZATIONS_FILE = "res/hidden_organizations.json"
IOTFINDER_VENDORS = load_json(IOTFINDER_VENDORS_FILE, "vendors_list")
IOTFINDER_VENDORS_FILE = "res/iotfinder_vendors.json"
RAW_INFO_FILE = "res/train_08.json"
TRAIN_RESULT_FILE = "res/train_08.json"
ALL_IOT_DOMAINS_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "all_iot_domains.json")
ALL_MERGED_IOT_DOMAINS_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "all_merged_iot_domains.json")
COMMON_VENDORS_LIST_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "common_vendors_list.json")
DEVICE_IP_DICT_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "device_ip_dict.json")
DOMAINS_BING_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "domains_bing.json")
DOMAINS_GOOGLE_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "domains_google.json")
DOMAINS_VENDORS_BING_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "domains_vendors_bing.json")
DOMAINS_VENDORS_GOOGLE_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "domains_vendors_google.json")
DOMAINS_VENDORS_INFO_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "domains_vendors_info.json")
DOMAINS_WHOIS_VENDOR_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "domains_whois_vendor.json")
IP_DEVICE_DICT_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "ip_device_dict.json")
LINUX_WHOIS_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "whois/linux_whois.json")
PYTHON_WHOIS_FILE = os.path.join(TEST_RESULT_FOLDER_NAME, "whois/python_whois.json")

# 实例化BaiduSpider
spider = BaiduSpider()


def baidu_spider():
    # 导入BaiduSpider
    from baiduspider import BaiduSpider
    from utils import store_json

    # 实例化BaiduSpider
    spider = BaiduSpider()

    # 搜索网页
    result = spider.search_web(query='roku b.canaryis.com')
    query_results_num = result["results"]["result"]

    store_json(result, 'crawl.res')


def get_folder_name():
    path = "E:/traffic_dataset\IoT SENTINEL Automated Device-Type Identification for Security Enforcement in IoT/IoT_Sentinel-master/captures_IoT_Sentinel/captures_IoT-Sentinel"
    dirs = os.listdir(path)
    for dir in dirs:
        print(dir)


def modify_filename():
    path = "E:/traffic_dataset/SoK Security Evaluation of Home-Based IoT Deployments/iot_traffic20180320/2018/03/20/"
    files = os.listdir(path)
    for filename in files:
        name, suffix = os.path.splitext(filename)
        new_name = os.path.join(path, name + ".pcap")
        old_name = os.path.join(path, filename)
        os.rename(old_name, new_name)


def my_sim1(vector_a, vector_b):
    """
    计算两个向量之间的相似度
    :param vector_a: 向量 a
    :param vector_b: 向量 b
    :return: sim
    """
    length = len(vector_a)
    sum = 0
    denom = 0
    for i in range(length):
        a = vector_a[i]
        b = vector_b[i]
        sum += a * b * np.log2(1 + min(a, b) / max(a, b))
        if b == 0:
            denom += a
        else:
            denom += a * b
    sim = sum / denom
    return sim


def my_sim2(vector_a, vector_b):
    """
    计算两个向量之间的相似度
    :param vector_a: 向量 a
    :param vector_b: 向量 b
    :return: sim
    """
    length = len(vector_a)
    sum = 0
    for i in range(length):
        a = vector_a[i]
        b = vector_b[i]
        if a != 0 and b != 0:
            sum += a * b * np.log2(1 + min(a, b) / max(a, b))
    sim = sum
    return sim


def output_vendor_xlsx(result):
    ip_pos_dict = dict()
    # 获取device_mapping.xlsx中的设备名以及ip列表
    devices_name = []
    ips = []
    for row in range(START_ROW, END_ROW + 1):
        devices_name.append(RAW_WS.cell(row, DEVICE_NAME_COL).value)
        ips.append(RAW_WS.cell(row, IP_COL).value)
        ip = RAW_WS.cell(row, IP_COL).value
        ip_pos_dict[ip] = (row, IP_COL)

    res_wb = Workbook()
    res_ws = res_wb.active
    # 复制device_mapping.xlsx中的设备名以及ip到新的vendor.xlsx中
    for row in range(START_ROW, END_ROW + 1):
        res_ws.cell(row, DEVICE_NAME_COL).value = devices_name[row - 1]
        res_ws.cell(row, IP_COL).value = ips[row - 1]
    # 填入vendor信息
    for ip, ip_info in result["ips_domains_tfidf"].items():
        vendors = ip_info["vendors_info"]
        for i in range(len(vendors)):
            res_ws.cell(ip_pos_dict[ip][0], VENDOR_START_COL + i).value = vendors[i]["vendor"]
    res_wb.save(RES_XLSX)


def fill_test_domain(train_domains, test_devices_fp):
    for train_domain in train_domains:
        for test_device_ip in test_devices_fp.keys():
            if train_domain not in test_devices_fp[test_device_ip].keys():
                test_devices_fp[test_device_ip][train_domain] = 0
    return test_devices_fp


def get_test_devices_fp(pkts, train_domains):
    devices_fp = {}
    devices_first_window = {}
    devices_last_window = {}
    devices_pks_last_window = {}
    start_time = pkts[0].time
    window_index = 0
    for i, pkt in enumerate(pkts):
        ip = pkt[sc.IP].dst
        domain = str(pkt[sc.DNS].qd.qname, encoding=ENCODING_METHOD)[:-1]
        window_index = math.floor((pkt.time - start_time) / WINDOW_SECONDS) + 1
        if ip in BLACK_IPS:
            continue
        if is_excluded_domain(domain, TOP_DOMAINS, EXCLUDED_DOMAINS_SUFFIX):
            continue
        domain = erase_protocol_prefix(domain)
        if ip not in devices_fp.keys():
            if domain in train_domains:
                devices_fp[ip] = {
                    domain: 1
                }
            else:
                devices_fp[ip] = {}
            devices_first_window[ip] = window_index
            devices_pks_last_window[ip] = {}
        elif domain in train_domains:
            if domain not in devices_fp[ip].keys():
                devices_fp[ip][domain] = 1
            elif window_index > devices_pks_last_window[ip][domain]:
                devices_fp[ip][domain] += 1
        devices_last_window[ip] = window_index
        devices_pks_last_window[ip][domain] = window_index
    devices_fp = fill_test_domain(train_domains, devices_fp)
    for ip, device_fp in devices_fp.items():
        for domain, counts in device_fp.items():
            devices_fp[ip][domain] = counts / (devices_last_window[ip] - devices_first_window[ip] + 1)
    return devices_fp, window_index


def cal_vectors_length(vectors):
    devices_vector_length = {}
    for ip, device_domain_tfidf_vector in vectors.items():
        domains_vector = []
        for domain_tfidf_vector in device_domain_tfidf_vector.values():
            domains_vector.append(domain_tfidf_vector)
        devices_vector_length[ip] = np.linalg.norm(domains_vector)
    return devices_vector_length


def cal_rate(true_devices_ip, test_devices_result, theta):
    tp = 0
    fp = 0
    tn = 0
    fn = 0
    for test_device_ip, score in test_devices_result:
        if test_device_ip in true_devices_ip:
            if score >= theta:
                tp += 1
            else:
                fn += 1
        else:
            if score >= theta:
                fp += 1
            else:
                tn += 1
    try:
        tpr = tp / (tp + fn)
        fpr = fp / (fp + tn)
        tnr = tn / (fp + tn)
        fnr = fn / (tp + fn)
    except Exception:
        return [tp, fp, tn, fn]
    return [tpr, fpr, tnr, fnr]


def old_similarity(tfidf_a, tfidf_b):
    vector_a = []
    vector_b = []
    for domain, tfidf in tfidf_a.items():
        vector_a.append(tfidf)
        vector_b.append(tfidf_b.get(domain, 0))

    return cos_sim(vector_a, vector_b)
    # return my_sim1(vector_a, vector_b)
    # return my_sim2(vector_a, vector_b)


def main():
    pass
    # 3.2 matching for EXAMINED_DEVICES_IP

    # test_pkts = rdpcap(TEST_PCAP_FILE)
    #
    # for examined_device_ip in EXAMINED_DEVICES_IP:
    #     print("\nUse %s as the train device:" % examined_device_ip)
    #     print("~" * 50)
    #     test_devices_fp, nt = get_test_devices_fp(test_pkts, train_devices_fp[examined_device_ip].keys())
    #     print("test_devices_fp: ", test_devices_fp)
    #     print("nt: ", nt)
    #
    #     test_domains = get_all_domains(test_devices_fp)
    #     print("test_domains: ", test_domains)
    #
    #     test_domains_devices_nums = get_domains_devices_nums(test_devices_fp)
    #     print("test_domains_devices_nums: ", test_domains_devices_nums)
    #
    #     test_domains_idf = get_domains_idf(test_domains)
    #     print("test_domains_idf: ", test_domains_idf)
    #
    #     # test_tfidf
    #     test_devices_domain_tfidf = tfidf(test_devices_fp, train_domains_idf)
    #     print("test_devices_domain_tfidf: ", test_devices_domain_tfidf)
    #
    #     # one_train_device_tfidf
    #     print("one_train_device_tfidf: ", train_devices_domains_tfidf[examined_device_ip])
    #
    #     # calculate the train and test vectors' tfidf vectorlength
    #     train_devices_tfidf_vector_length = cal_vectors_length(train_devices_domains_tfidf)
    #     test_devices_tfidf_vector_length = cal_vectors_length(test_devices_domain_tfidf)
    #
    #     # compute cosine similarity
    #     devices_cosine_similarity = {}
    #     for test_devices_ip, test_devices_tfidf in test_devices_domain_tfidf.items():
    #         cosine_similarity = old_similarity(train_devices_domains_tfidf[examined_device_ip],
    #                                                         test_devices_tfidf)
    #         if cosine_similarity != 0:
    #             cosine_similarity = cosine_similarity / (train_devices_tfidf_vector_length[examined_device_ip]
    #                                                      * test_devices_tfidf_vector_length[test_devices_ip])
    #         devices_cosine_similarity[test_devices_ip] = cosine_similarity
    #
    #     # all test devices result
    #     print("ip   score")
    #     test_devices_result = dict2sorted_list(devices_cosine_similarity)
    #     for ip, score in test_devices_result:
    #         print(ip, score)
    #
    #     # calculate the rate
    #     rate = cal_rate([examined_device_ip], test_devices_result, THETA)
    #     print("tpr fpr tnr fnr: ", rate)


# def filterHtmlTag(htmlstr):
#     '''
#     过滤html中的标签
#     '''
#     # 兼容换行
#     s = htmlstr.replace('\r\n', '\n')
#     s = htmlstr.replace('\r', '\n')
#
#     # 规则
#     re_cdata = re.compile('//<!\[CDATA\[[^>]*//\]\]>', re.I)  # 匹配CDATA
#     re_script = re.compile('<\s*script[^>]*>[\S\s]*?<\s*/\s*script\s*>', re.I)  # script
#     re_style = re.compile('<\s*style[^>]*>[\S\s]*?<\s*/\s*style\s*>', re.I)  # style
#     re_br = re.compile('<br\\s*?\/??>', re.I)  # br标签换行
#     re_p = re.compile('<\/p>', re.I)  # p标签换行
#     re_h = re.compile('<[\!|/]?\w+[^>]*>', re.I)  # HTML标签
#     re_comment = re.compile('<!--[^>]*-->')  # HTML注释
#     re_hendstr = re.compile('^\s*|\s*$')  # 头尾空白字符
#     re_lineblank = re.compile('[\t\f\v ]*')  # 空白字符
#     re_linenum = re.compile('\n+')  # 连续换行保留1个
#
#     # 处理
#     s = re_cdata.sub('', s)  # 去CDATA
#     s = re_script.sub('', s)  # 去script
#     s = re_style.sub('', s)  # 去style
#     s = re_br.sub('\n', s)  # br标签换行
#     s = re_p.sub('\n', s)  # p标签换行
#     s = re_h.sub('', s)  # 去HTML标签
#     s = re_comment.sub('', s)  # 去HTML注释
#     s = re_lineblank.sub('', s)  # 去空白字符
#     s = re_linenum.sub('\n', s)  # 连续换行保留1个
#     s = re_hendstr.sub('', s)  # 去头尾空白字符
#
#     # 替换实体
#     s = replaceCharEntity(s)
#
#     return s
#
#
# def replaceCharEntity(htmlStr):
#     '''
#       替换html中常用的字符实体
#       使用正常的字符替换html中特殊的字符实体
#       可以添加新的字符实体到CHAR_ENTITIES 中
#       CHAR_ENTITIES是一个字典前面是特殊字符实体  后面是其对应的正常字符
#       :param htmlStr:
#       '''
#     CHAR_ENTITIES = {'nbsp': ' ', '160': ' ',
#                      'lt': '<', '60': '<',
#                      'gt': '>', '62': '>',
#                      'amp': '&', '38': '&',
#                      'quot': '"', '34': '"', }
#     re_charEntity = re.compile(r'&#?(?P<name>\w+);')
#     sz = re_charEntity.search(htmlStr)
#     while sz:
#         entity = sz.group()  # entity全称，如>
#         key = sz.group('name')  # 去除&;后的字符如（" "--->key = "nbsp"）    去除&;后entity,如>为gt
#         try:
#             htmlStr = re_charEntity.sub(CHAR_ENTITIES[key], htmlStr, 1)
#             sz = re_charEntity.search(htmlStr)
#         except KeyError:
#             # 以空串代替
#             htmlStr = re_charEntity.sub('', htmlStr, 1)
#             sz = re_charEntity.search(htmlStr)
#     return htmlStr


def get_whois(domain):
    logger.info("whois api for: {}".format(domain))
    url = WHOIS_RAW_URL.format(WHOIS_API_KEY, domain, WHOIS_OUTPUT_FORMAT)
    r = requests.get(url)
    return json.loads(r.text)


def get_whois_vendor(domain):
    # 隐藏的组织信息，可以用这个来区分一个域名的组织信息是否是被隐藏的
    hidden_organizations = load_json(HIDDEN_ORGANIZATIONS_FILE, "organizations")
    registrant_organization = None
    # if domain in domains_whois_info.keys():
    #     whois_info = domains_whois_info[domain]
    # else:
    #     whois_info = get_whois(domain)
    #     domains_whois_info[domain] = whois_info

    whois_info = get_domain_whois(domain)
    registrant_organization = whois_info["WhoisRecord"]["registrant"]["organization"]
    if registrant_organization is None or registrant_organization in hidden_organizations:
        return None
    else:
        return registrant_organization.rsplit(" ", maxsplit=1)[0].split(",")[0].lower()


def get_domain_whois(domain):
    """
    获取domain的whois信息
    :param domain:
    :return:
    """
    logger.info("whois mongodb: {}".format(domain))
    whois_info = WHOIS_COL.find_one({"domain": domain})
    if whois_info is None:  # 如果MongoDB中没有该domain的whois信息
        whois_record = get_whois(domain)  # 使用whois api获取domain信息
        whois_info = {"domain": domain}
        whois_info.update(whois_record)
        WHOIS_COL.insert_one(whois_info)  # 将获取的信息写入到MongoDB中
    return whois_info


# 使用bing搜索，返回查询结果数
def bing(key_word):
    # 搜索网页
    result = spider.search_web(query=key_word)
    print("-" * 100)
    print("search: " + key_word)
    print("-" * 100)
    if len(result["results"]) != 0:
        query_results_num = result["results"][0]["result"]
    else:
        query_results_num = 0
    return query_results_num


# 返回一个列表，列表元素是二元组("vendor", {"support": support, "confidence": confidence, "score": score})，按score由大到小排序
def get_all_bing(domains, vendors):
    """
    利用搜索引擎，搜索每个域名的查询数，以及和vendor关联时的查询数
    :param domains:
    :param vendors:
    :return:
    """
    domains_bing = load_json(DOMAINS_BING_FILE)
    domains_vendors_bing = load_json(DOMAINS_VENDORS_BING_FILE)
    try:
        for domain in domains:
            if domain not in domains_bing.keys():
                domains_bing[domain] = bing(domain)
            if domain not in domains_vendors_bing.keys():
                domains_vendors_bing[domain] = dict()
            for vendor in vendors:
                if vendor not in domains_vendors_bing[domain].keys():
                    domains_vendors_bing[domain][vendor] = bing(domain + ' ' + vendor)
    finally:
        store_json(domains_bing, DOMAINS_BING_FILE)
        store_json(domains_vendors_bing, DOMAINS_VENDORS_BING_FILE)


# 返回字典{"vendor", {"support": support, "confidence": confidence, "score": score}}，按score由大到小排序
# def get_guessed_vendor(domain, vendors_list):
#     if domain in domains_bing.keys():
#         domain_bing = domains_bing[domain]
#     else:
#         domain_bing = bing(domain)
#         domains_bing[domain] = domain_bing
#     domain_vendors_info = dict()  # {"*vendor": {"support": support, "confidence": confidence, "score": score}}
#     domain_vendors_bing_dict = dict()  # {"*vendor": domain_vendor_bing}
#     domain_vendors_bing_sum = 1
#     for vendor in vendors_list:
#         if domain in domains_vendors_bing.keys() and vendor in domains_vendors_bing[domain].keys():
#             domain_vendor_bing = domains_vendors_bing[domain][vendor]
#         else:
#             domain_vendor_bing = bing(domain + " " + vendor)
#             # fill domains_vendors_bing
#             if domain not in domains_vendors_bing.keys():
#                 domains_vendors_bing[domain] = dict()
#             domains_vendors_bing[domain][vendor] = domain_vendor_bing
#         domain_vendors_bing_dict[vendor] = domain_vendor_bing
#         domain_vendors_bing_sum += domain_vendor_bing
#     domain_vendor_support_threshold = domain_vendors_bing_sum * support_threshold
#     domain_vendor_confidence_threshold = domain_bing * confidence_threshold
#     domain_vendors_info_list = list()
#     for vendor in vendors_list:
#         # if domain_vendors_bing_dict["vendor"] >= max(domain_vendor_support_threshold, domain_vendor_confidence_threshold):
#         domain_vendors_info[vendor] = {"support": domain_vendors_bing_dict[vendor] / domain_vendors_bing_sum,
#                                        "confidence": domain_vendors_bing_dict[vendor] / domain_bing
#                                        }
#         domain_vendors_info[vendor]["score"] = adjust_score(vendor, COMMON_VENDORS,
#                                                             domain_vendors_info[vendor]["support"],
#                                                             domain_vendors_info[vendor]["confidence"])
#     # domain_vendors_info_list = dict2sorted_list(domain_vendors_info, "score")
#     # domain_vendors_info = dict()
#     # for vendor, vendor_info in domain_vendors_info_list:
#     #     domain_vendors_info[vendor] = vendor_info
#     return get_sorted_dict(domain_vendors_info, compared_target="value", compared_name="score")


def get_great_vendor_by_google(domain, vendors_list, domains_vendors_google_relation):
    vendors_info = get_guessed_vendor(domain, vendors_list)
    domains_vendors_google_relation[domain] = vendors_info
    great_vendor = get_key_by_max_value(vendors_info, "score")
    return great_vendor


def get_all_merged_iot_domains():
    all_iot_domains = load_json(ALL_IOT_DOMAINS_FILE)
    all_merged_iot_domains = load_json(ALL_MERGED_IOT_DOMAINS_FILE)
    merged_domains = get_merged_domains(all_iot_domains)
    all_merged_iot_domains["all_merged_iot_domains"] = merged_domains
    store_json(all_merged_iot_domains, ALL_MERGED_IOT_DOMAINS_FILE)


def google_with_multiple_threads():
    all_merged_iot_domains = load_json(ALL_MERGED_IOT_DOMAINS_FILE, "all_merged_iot_domains")
    try:
        for i in range(28, 29):
            _thread.start_new_thread(get_all_google, (all_merged_iot_domains[i:i + 1], IOTFINDER_VENDORS))
    except:
        print("返回无效结果")
    while 1:
        pass


def get_devices_mixed_domains_tfidf():
    """
    将设备的域名进行聚合  general_domains中的b.a.com 和 c.a.com 聚合成a.com
    device -> classified_domains, general_domains
    :return:
    """
    devices_domains_tfidf = load_json(DEVICES_DOMAINS_TFIDF_FILE)
    # confusion_domains = load_json(CONFUSION_DOMAINS_FILE)
    devices_mixed_domains_tfidf = dict()
    for device, domains_tfidf in devices_domains_tfidf.items():
        devices_mixed_domains_tfidf[device] = dict()
        devices_mixed_domains_tfidf[device]["classified_domains"] = dict()
        device_right_patterns_domains = dict()
        for domain, tfidf in domains_tfidf.items():
            sub_domain, second_domain, suffix = tldextract.extract(domain)
            right_pattern = second_domain + '.' + suffix
            # if right_pattern.lower() in confusion_domains:
            #     if right_pattern.lower() not in device_right_patterns_domains.keys():
            #         device_right_patterns_domains[right_pattern.lower()] = list()
            #     device_right_patterns_domains[right_pattern.lower()].append(domain)
            # else:
            if right_pattern not in device_right_patterns_domains.keys():
                device_right_patterns_domains[right_pattern] = list()
            device_right_patterns_domains[right_pattern].append(domain)
        for right_pattern in device_right_patterns_domains.keys():
            if len(device_right_patterns_domains[right_pattern]) >= 2:
                devices_mixed_domains_tfidf[device]["classified_domains"][right_pattern] = \
                    device_right_patterns_domains[right_pattern]
        devices_mixed_domains_tfidf[device]["general_domains"] = domains_tfidf
    store_json(devices_mixed_domains_tfidf, DEVICES_MIXED_DOMAINS_TFIDF_FILE)


def get_train_result(windows_num, train_ips_domains_tfidf):
    """
    获取训练数据的结果
    :param windows_num:  窗口数
    :param train_ips_domains_tfidf: 训练集中每个ip的域名tfidf信息
    :return:
    """

    def ips_info_to_devices_info(train_ips_domains_info):
        """
        内置函数，由于训练集中每个ip对应一个设备，将ip的信息转为device的信息，并附加上ip
        :param train_ips_domains_info: 每个ip的域名tfidf信息
        :return: 每个设备每个域名的tfidf信息
        """
        device_to_ip = load_json(DEVICE_TO_IP_FILE)
        ip_to_device = dict(zip(device_to_ip.values(), device_to_ip.keys()))
        result = dict()
        for ip, domains_info in train_ips_domains_info.items():
            device = ip_to_device[ip]
            result[device] = dict()
            result[device]["ip"] = ip
            domains_list = dict2sorted_list(domains_info)
            domains_info = dict()
            for domain_info in domains_list:
                domains_info[domain_info[0]] = domain_info[1]
            result[device]["domains_tfidf"] = domains_info
        return result

    train_result = dict()
    train_result["pcap"] = TRAIN_PCAP_FILE
    train_result["window_seconds"] = WINDOW_SECONDS
    train_result["windows_num"] = windows_num
    train_result["devices_info"] = ips_info_to_devices_info(train_ips_domains_tfidf)
    store_json(train_result, TRAIN_RESULT_FILE)


def get_train_device_to_ip():
    """
    获取训练集中device和ip的映射关系  device -> ip
    :return:
    """
    devices_info = load_json(DEVICES_INFO_FILE)
    device_to_ip = dict()
    for device, device_info in devices_info.items():
        device_to_ip[device] = device_info[DEVICES_INFO_IP_POS]
    store_json(device_to_ip, DEVICE_TO_IP_FILE)


def get_devices_great_domains_to_devices():
    """
    这里从每个device的domain集M中挑出得分最高的2个domain，形成{"*domain": [device1, ...]}
    在测试ip的domain集N中每个domain，查询M中键为domain的设备，得到的设备集合记为D，
    将测试ip和D中每个设备d进行tfidf匹配，设定阈值为θ，得到满足相似度大于θ的猜测设备集G
    :return:
    """
    devices_domain_tfidf = load_json(DEVICES_DOMAINS_TFIDF_FILE)
    devices_great_domains_to_devices = dict()
    # 各个IoT设备最显著的2个domain组成domain集
    for device, domains_tfidf_info in devices_domain_tfidf.items():
        great_domain_count = 0
        for domain in domains_tfidf_info.keys():
            great_domain_count += 1
            if domain not in devices_great_domains_to_devices.keys():
                devices_great_domains_to_devices[domain] = [device]
            else:
                devices_great_domains_to_devices[domain].append(device)
            if great_domain_count == 2:  # 挑得分最高的两个
                break
    store_json(devices_great_domains_to_devices, DEVICES_GREAT_DOMAINS_TO_DEVICES_FILE)


def detect_domain():
    user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'
    headers = {'User-Agent': user_agent}
    devices_domains_tfidf = load_json(DEVICES_DOMAINS_TFIDF_FILE)
    devices_domains_detection = dict()
    for device, domains_tfidf in devices_domains_tfidf.items():
        if device == "GoogleHomeMini":
            break
        devices_domains_detection[device] = dict()
        for domain in domains_tfidf.keys():
            try:
                full_url = "http://" + domain
                r_text = requests.get(full_url, headers=headers).text
                devices_domains_detection[device][domain] = {
                    "text": r_text,
                    "bytes": len(r_text)
                }
            except Exception:
                devices_domains_detection[device][domain] = {
                    "text": None,
                    "bytes": 0
                }
            logger.info("device: {device}, domain: {domain}, detection: {detection}"
                        .format(device=device, domain=domain, detection=devices_domains_detection[device][domain]))
    store_json(devices_domains_detection, DEVICES_DOMAINS_DETECTION_FILE)


def get_devices_domains_tfidf_vector_length():
    """
    获取训练集中每个设备的tf-idf向量长度，保存到文件中，避免计算相似度时重复计算
    :return:
    """
    devices_domains_tfidf = load_json(DEVICES_DOMAINS_TFIDF_FILE)
    devices_domains_tfidf_vector_length = dict()
    for device, domains_tfidf in devices_domains_tfidf.items():
        devices_domains_tfidf_vector_length[device] = cal_tfidf_vectors_length(domains_tfidf)
    store_json(devices_domains_tfidf_vector_length, DEVICES_DOMAINS_TFIDF_VECTOR_LENGTH_FILE)


def get_devices_domains_tfidf():
    """
    获取每个设备每个域名的tfidf，并且按tfidf值从大到小进行排序
    :return:
    """
    devices_domains_tfidf = dict()
    devices_info = load_json(TRAIN_RESULT_FILE, "devices_info")
    for device, info in devices_info.items():
        devices_domains_tfidf[device] = info["domains_tfidf"]
    store_json(devices_domains_tfidf, DEVICES_DOMAINS_TFIDF_FILE)


def get_organization_from_whois():
    """
    从已有的domains_whois_info中，获取domain对应的组织信息
    :return:
    """
    domains_whois_info = load_json(DOMAINS_WHOIS_INFO_FILE)
    domains_whois_vendor = load_json(DOMAINS_WHOIS_VENDOR_FILE)
    for domain in domains_whois_info.keys():
        try:
            vendor = domains_whois_info[domain]["WhoisRecord"]["registrant"]["organization"]
        except:
            vendor = None
        domains_whois_vendor[domain] = vendor
    store_json(domains_whois_vendor, DOMAINS_WHOIS_VENDOR_FILE)


def get_all_iot_domains_idf(train_domains_devices_nums, iot_clients_num):
    """
    生成每个domain的idf，本来应该是用全球iot域名知识库，但这个获取不到，暂时用目前数据集有的
    :param train_domains_devices_nums: 每个域名在多少个iot设备中出现过
    :param iot_clients_num: iot设备的数目
    :return:
    """
    all_iot_domains_idf = dict()
    for domain, count in train_domains_devices_nums.items():
        all_iot_domains_idf[domain] = math.log((1 + (iot_clients_num / (1 + count))), 2)
    store_json(all_iot_domains_idf, ALL_IOT_DOMAINS_IDF_FILE)


def use_python_whois():
    supported_tld = ['com', 'at', 'uk', 'pl', 'be', 'biz', 'br', 'ca', 'co', 'jp', 'co_jp', 'cz', 'de', 'eu', 'fr',
                     'info', 'io', 'it', 'ru', 'lv', 'me', 'mx', 'name', 'net', 'nz', 'org', 'ru_rf', 'sh', 'us']
    # test_ips_info = load_json(TEST_IPS_INFO_FILE, "ips_domains_tfidf")
    test_ips_info = load_json("res/lazy.json", "ips_domains_tfidf")
    python_whois = dict()
    search_num = 0
    for ip, domains_tfidf in test_ips_info.items():
        if ip == '192.168.12.250':
            continue
        python_whois[ip] = dict()
        merged_domains = get_merged_domains(domains_tfidf.keys())
        for merged_domain in merged_domains:
            sub_domain, second_domain, suffix = tldextract.extract(merged_domain)
            if suffix not in supported_tld:
                continue
            if search_num == 10:
                store_json(python_whois, PYTHON_WHOIS_FILE)
                return
            search_num += 1
            domain_dict = whois.query(merged_domain).__dict__
            domain_dict["name_servers"] = list(domain_dict["name_servers"])
            python_whois[ip][merged_domain] = domain_dict
            time.sleep(10)
    store_json(python_whois, PYTHON_WHOIS_FILE)


def get_top_domains(filename):
    """
    从csv中获取数据并存储为json
    :param filename: csv
    :return:
    """
    top_domains = []
    with open(filename) as f:
        f_csv = csv.reader(f)
        i = 0
        for row in f_csv:
            top_domains.append(row[1])
            i += 1
            if i == 100:
                break
    store_json(top_domains, TOP_DOMAINS_FILE)


def adjust_score(vendor, common_vendors, support, confidence):
    """
    对于domain下每个vendor的得分，还需要判断该vendor是否是常见的vendor乘一个系数来调整
    :param vendor:
    :param common_vendors:
    :param support:
    :param confidence:
    :return:
    """
    if vendor in common_vendors:
        return 1 * support * confidence
    else:
        return 1 * support * confidence
    return filtered_domains_tfidf_2


def get_guessed_vendor(domain, vendors_list):
    """
    获取一个domain下所有vendor的搜索结果的三个指标
    :param domain: 某个域名
    :param vendors_list: 所有要搜索的vendor
    :return: {"vendor", {"support": support, "confidence": confidence, "score": score}}，多个vendor，按score由大到小排序
    """
    domain_search_num = get_word_google_search_num(domain)
    vendors_measure = dict()  # {"*vendor": {"support": support, "confidence": confidence, "score": score}}
    domain_vendors_search_num_dict = dict()  # {"*vendor": domain_vendor_search_num}
    domain_vendors_search_num_sum = 1
    for vendor in vendors_list:
        search_num = get_word_google_search_num(domain + " " + vendor)
        domain_vendors_search_num_dict[vendor] = search_num
        domain_vendors_search_num_sum += search_num
    domain_vendor_support_threshold = domain_vendors_search_num_sum * support_threshold
    domain_vendor_confidence_threshold = domain_search_num * confidence_threshold
    for vendor in vendors_list:
        # if domain_vendors_search_num_dict["vendor"] >= max(domain_vendor_support_threshold, domain_vendor_confidence_threshold):
        vendors_measure[vendor] = {"support": domain_vendors_search_num_dict[vendor] / domain_vendors_search_num_sum,
                                   "confidence": domain_vendors_search_num_dict[vendor] / domain_search_num
                                   }
        vendors_measure[vendor]["score"] = adjust_score(vendor, COMMON_VENDORS,
                                                        vendors_measure[vendor]["support"],
                                                        vendors_measure[vendor]["confidence"])
    return get_sorted_dict(vendors_measure, compared_target="value", compared_name="score")


def draw_theta():
    """
    根据各个阈值对应的分数，画出曲线
    :return:
    """
    all_thetas_performance = load_json(ALL_THETAS_PERFORMANCE_FILE)
    thetas = list()
    scores = list()
    for theta_str in all_thetas_performance.keys():
        thetas.append(float(theta_str[:THETA_STR_LENGTH]))
        # scores.append(all_thetas_performance[theta]["precision"])
        scores.append(all_thetas_performance[theta_str]["recall"])
        # scores.append(all_thetas_performance[theta]["score"])
    plt.title("Roberto's 12-hour DNS dataset")
    plt.xlabel("threshold")
    # plt.ylabel("precision")
    plt.ylabel("recall")
    # plt.ylabel("F2 score")
    max_score = max(scores)
    max_index = scores.index(max_score)
    max_theta = thetas[max_index]
    show_max = "(" + str(max_theta) + ", " + str(max_score)[:5] + ")"
    plt.annotate(show_max, xytext=(max_theta, max_score), xy=(max_theta, max_score), )
    plt.plot(thetas, scores, 'r-o')
    plt.show()


def delete_trash_suffix_urls():
    """
    删除MongoDB中的垃圾URL
    :return:
    """
    for url in URLS_COL.find({"url": {"$regex": ".*\.[\w]{2,}$"}}):
        suffix = url["url"].rsplit(".", 1)[1]
        if suffix in MONGODB_NOT_DOWNLOAD_SUFFIX:
            print(url["url"])
            URLS_COL.delete_one({"url": url["url"]})


def delete_urls_proxy_error():
    """
    删除由于代理错误导致异常的url
    :return:
    """
    for url in URLS_COL.find({"exception": {"$regex": "^ProxyError"}}):
        print(url["url"])
        URLS_COL.delete_one({"url": url["url"]})


def extend_google_related():
    """
    扩展原GOOGLE_COL中page_info中的related_searches，谷歌搜索页面可能有相关搜索的条目，类似于百度知道
    :return:
    """
    for google in GOOGLE_COL.find():
        query_info = google["query_info"]
        page_info = google["page_info"]
        page_info["related_searches"] = None
        GOOGLE_COL.update_one({"query_info": query_info}, {"$set": {"page_info": page_info}})


def whois2mongo():
    """
    将文件形式的whois存储到MongoDB中
    :return:
    """
    domains_whois_info = load_json(DOMAINS_WHOIS_INFO_FILE)

    client = pymongo.MongoClient("mongodb://localhost:27017/")
    iot_db = client["iot"]
    whois_col = iot_db["whois"]

    for domain, whois_record in domains_whois_info.items():
        inserted_dict = {"domain": domain}
        inserted_dict.update(whois_record)
        whois_col.insert_one(inserted_dict)


def google2mongo():
    """
    将文件形式的google存储到MongoDB中
    :return:
    """
    domains_vendors_google = load_json(DOMAINS_VENDORS_GOOGLE_FILE)

    client = pymongo.MongoClient("mongodb://localhost:27017/")
    iot_db = client["iot"]
    google_col = iot_db["old_google"]

    for domain, vendors_search_info in domains_vendors_google.items():
        for vendor, search_info in vendors_search_info.items():
            inserted_dict = {
                "search_word": domain + " " + vendor,
                "search_result": search_info
            }
            google_col.insert_one(inserted_dict)


def get_domain_whois_by_web():
    """
    用网站来获取whois信息
    :return:
    """
    user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'
    headers = {'User-Agent': user_agent}
    # url = "https://www.whois.com/whois/xboxlive.com"
    url = "https://www.google.com"
    proxies = {
        "http": '192.168.1.101:7890',
        "https": '192.168.1.101:7890'
    }
    # url = "https://" + url
    r = requests.get(url, headers=headers, proxies=proxies)
    print(r.text)
    print(len(r.text))


def cal_cycle():
    """
    猜测周期
    :return:
    """
    tolerance_time = 2
    time_list1 = [0, 300, 601, 903]
    time_list2 = [50, 351, 653, 954]
    # time_list = [0, 50, 300, 351, 601, 653, 903, 954]
    time_list = list(set(time_list1) | set(time_list2))
    time_list.sort()
    print(time_list)
    delta_time_list = list()
    for i in range(1, len(time_list)):
        prev_time = time_list[i - 1]
        cur_time = time_list[i]
        delta_time_list.append(cur_time - prev_time)
    print(delta_time_list)
    guessed_cycles = list()
    guessed_cycles.append(time_list[1] - time_list[0])
    guessed_cycles.append(time_list[2] - time_list[0])
    start_time = time_list[0]
    end_time = time_list[-1]
    is_iot_domain = False
    for guessed_cycle in guessed_cycles:
        cycle_count = 0
        for i, dns_time in enumerate(time_list):
            if guessed_cycle * cycle_count - tolerance_time <= dns_time <= guessed_cycle * cycle_count + tolerance_time:
                cycle_count += 1
        if cycle_count - 1 >= (end_time - start_time) / guessed_cycle * 0.6:
            is_iot_domain = True
            break
    return is_iot_domain


# def measure_performance(test_ips_possible_devices_similarity, test_ips_devices_info, theta):
#     test_ips_nums = len(test_ips_possible_devices_similarity)
#     ips_precision = 0
#     ips_recall = 0
#     for test_ip, possible_devices_info in test_ips_possible_devices_similarity.items():
#         test_ip_devices_info_num = len(test_ips_devices_info[test_ip])
#         tp = 0
#         fp = 0
#         for possible_device, similarity in possible_devices_info.items():
#             if similarity >= theta:
#                 tp_flag = False
#                 for device_info in test_ips_devices_info[test_ip]:
#                     if possible_device == device_info["device"]:
#                         tp += 1
#                         tp_flag = True
#                         break
#                 if not tp_flag:
#                     fp += 1
#         fn = test_ip_devices_info_num - tp
#         if tp + fp == 0:  # 推测该ip没有iot设备
#             if test_ip_devices_info_num == 0:  # 该ip实际上没有任何iot设备
#                 ip_precision = 1
#             else:  # 该ip实际上有iot设备
#                 ip_precision = 0
#         else:  # 推测该ip有iot设备
#             ip_precision = tp / (tp + fp)  # 在预测为真的样例中，实际为真的概率
#         if tp + fn == 0:  # 表示该ip实际上没有任何iot设备，则默认召回率为1
#             ip_recall = 1
#         else:
#             ip_recall = tp / (tp + fn)  # 在实际为真的样例中，预测为真的概率
#         ips_precision += ip_precision
#         ips_recall += ip_recall
#     precision_average = ips_precision / test_ips_nums
#     recall = ips_recall / test_ips_nums
#     if precision_average + recall == 0:
#         score = 0
#     else:
#         score = (2 * precision_average * recall) / (precision_average + recall)
#     return precision_average, recall, score


def try_theta_for_possible_devices():
    """
    尝试判别的theta，若ip和某设备的tfidf相似度大于该阈值，视为ip存在该设备
    :return:
    """
    test_ips_possible_devices_similarity = load_json(TEST_IPS_POSSIBLE_DEVICES_SIMILARITY_FILE)
    test_ips_devices_info = load_json(TEST_IPS_DEVICES_INFO_FILE)
    all_thetas_performance = dict()
    for theta in np.arange(THETA_LOW, THETA_HIGH + THETA_STEP, THETA_STEP):
        precision, recall, f_05, f_1, f_2 = measure_performance(test_ips_possible_devices_similarity,
                                                                test_ips_devices_info, theta)
        all_thetas_performance[theta] = {
            "theta": theta,
            "precision": precision,
            "recall": recall,
            "F0.5": f_05,
            "F1": f_1,
            "F2": f_2,
        }
    store_json(all_thetas_performance, ALL_THETAS_PERFORMANCE_FILE)


def get_test_ips_performance():
    """
    得到所有ip的表现
    :return:
    """
    devices_threshold = get_mongodb_devices_threshold()
    test_ips_possible_devices_similarity = load_json(TEST_IPS_POSSIBLE_DEVICES_SIMILARITY_FILE)
    test_ips_devices_info = load_json(TEST_IPS_DEVICES_INFO_FILE)
    test_ips_performance = dict()
    test_ips_performance["all_info"] = dict()
    test_ips_performance["ips_info"] = dict()
    precision, recall, f_05, f_1, f_2 = measure_performance(test_ips_possible_devices_similarity,
                                                            test_ips_devices_info, theta)
    test_ips_performance["all_info"] = {
        "theta": theta,
        "precision": precision,
        "recall": recall,
        "F0.5": f_05,
        "F1": f_1,
        "F2": f_2,
    }
    store_json(test_ips_performance, TEST_IPS_PERFORMANCE_FILE)


def main():
    pass

    # # 1.1 从pcap包中获取源数据
    # ips_windows_num, ips_domains_windows_num = get_train_info(TRAIN_PCAP_FILE, EXCLUDED_DOMAINS_SUFFIX,
    #                                                                            train_ips=TRAIN_IPS)
    # print("train_ips_domains_frequency: ", train_ips_domains_frequency)
    # print("windows_num: ", windows_num)
    #
    # # 每个域名在其对应的设备的出现次数，每个域名的
    # # 每个域名的知识库保存在MongoDB中
    #
    # # 获取每个域名对应多少个设备
    # train_domains_devices_nums = get_domains_devices_nums(train_ips_domains_frequency)
    # print("train_domains_devices_nums: ", train_domains_devices_nums)
    #
    # # 获取所有的iot设备数
    # iot_clients_num = len(train_ips_domains_frequency)
    #
    # # 1.2 生成每个domain的idf
    # get_all_iot_domains_idf(train_domains_devices_nums, iot_clients_num)
    #
    # # 2.1 获取每个ip下每个domain的tfidf
    # all_iot_domains_idf = load_json(ALL_IOT_DOMAINS_IDF_FILE)
    # train_ips_domains_tfidf = get_ips_domains_tfidf(train_ips_domains_frequency, all_iot_domains_idf)
    # print("train_ips_domains_tfidf: ", train_ips_domains_tfidf)
    #
    # # 2.2 存储训练数据
    # get_train_result(windows_num, train_ips_domains_tfidf)

    # all_iot_domains = load_json(ALL_IOT_DOMAINS_FILE, "all_iot_domains")
    # iotfinder_vendors = load_json(IOTFINDER_VENDORS_FILE)
    # get_all_bing(all_iot_domains[:800], iotfinder_vendors)  # 使用baidu spider进行启发式搜索
    # get_devices_domains_tfidf()  # 获取每个设备每个域名的tfidf
    # get_devices_domains_tfidf_vector_length()  # 获取训练集中每个设备的tf-idf向量长度，保存到文件中，避免计算相似度时重复计算
    # get_devices_great_domains_to_devices()  # 每个设备挑2个最佳域名，得到 domain -> devices
    # get_devices_mixed_domains_tfidf()  # device -> classified_domains, general_domains
    # get_organization_from_whois()  # 从已有的domains_whois_info中，获取domain对应的组织信息


if __name__ == '__main__':
    main()
